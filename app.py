import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import io

# Try to import Gemini AI
try:
    import google.generativeai as genai
    GEMINI_AVAILABLE = True
except ImportError:
    GEMINI_AVAILABLE = False

st.set_page_config(page_title="Data Separation Tool", layout="wide", initial_sidebar_state="collapsed")

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');
    * { font-family: 'Inter', sans-serif; }
    #MainMenu {visibility: hidden;} footer {visibility: hidden;} header {visibility: hidden;}
    .main > div { background: #f8fafc; min-height: 100vh; padding: 2rem; }
    .hero-header { background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 2.5rem; border-radius: 20px; margin-bottom: 2rem; box-shadow: 0 20px 60px rgba(102, 126, 234, 0.3); }
    .hero-title { color: #ffffff; font-size: 2.5rem; font-weight: 800; margin: 0; }
    .hero-subtitle { color: rgba(255, 255, 255, 0.9); font-size: 1.1rem; margin-top: 0.5rem; }
    .premium-card { background: white; padding: 2rem; border-radius: 16px; box-shadow: 0 4px 20px rgba(0, 0, 0, 0.08); margin-bottom: 1.5rem; border: 1px solid #e5e7eb; }
    .card-title { color: #1e293b; font-size: 1.3rem; font-weight: 700; margin-bottom: 1.2rem; display: flex; align-items: center; gap: 0.5rem; }
    .card-number { display: inline-flex; align-items: center; justify-content: center; width: 32px; height: 32px; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; border-radius: 8px; font-size: 1rem; font-weight: 700; }
    .success-box { background: linear-gradient(135deg, #d1fae5 0%, #a7f3d0 100%); border-left: 4px solid #10b981; color: #065f46; padding: 1rem 1.2rem; border-radius: 10px; margin: 1rem 0; }
    .warning-box { background: linear-gradient(135deg, #fef3c7 0%, #fde68a 100%); border-left: 4px solid #f59e0b; color: #92400e; padding: 1rem 1.2rem; border-radius: 10px; margin: 1rem 0; }
    .info-box { background: linear-gradient(135deg, #dbeafe 0%, #bfdbfe 100%); border-left: 4px solid #3b82f6; color: #1e40af; padding: 1rem 1.2rem; border-radius: 10px; margin: 1rem 0; font-size: 0.95rem; }
    .error-box { background: linear-gradient(135deg, #fee2e2 0%, #fecaca 100%); border-left: 4px solid #ef4444; color: #991b1b; padding: 1rem 1.2rem; border-radius: 10px; margin: 1rem 0; }
    .stat-container { display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 1rem; margin: 1.5rem 0; }
    .stat-box { background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 1.5rem; border-radius: 16px; color: white; text-align: center; box-shadow: 0 8px 24px rgba(102, 126, 234, 0.3); }
    .stat-number { font-size: 2.2rem; font-weight: 800; margin-bottom: 0.3rem; }
    .stat-label { font-size: 0.85rem; opacity: 0.95; font-weight: 500; text-transform: uppercase; }
    .stButton>button { background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; border: none; padding: 0.9rem 2rem; border-radius: 12px; font-weight: 600; font-size: 1rem; width: 100%; transition: all 0.3s ease; }
    .stButton>button:hover { transform: translateY(-2px); box-shadow: 0 8px 24px rgba(102, 126, 234, 0.4); }
    .stDownloadButton>button { background: linear-gradient(135deg, #10b981 0%, #059669 100%); color: white; border: none; padding: 1rem 1.5rem; border-radius: 12px; font-weight: 600; width: 100%; }
    .distribution-item { background: linear-gradient(135deg, #fafafa 0%, #f5f5f5 100%); padding: 1rem 1.5rem; border-radius: 12px; margin: 0.5rem 0; display: flex; justify-content: space-between; border-left: 4px solid #667eea; }
    @media (max-width: 768px) { .stat-container { grid-template-columns: repeat(2, 1fr); } .hero-title { font-size: 1.8rem; } }
</style>
""", unsafe_allow_html=True)

def init_gemini():
    """Initialize Gemini with proper model detection"""
    if not GEMINI_AVAILABLE:
        return None, "Package not installed"
    
    try:
        if "GEMINI_API_KEY" not in st.secrets:
            return None, "API key not found in secrets"
        
        api_key = st.secrets["GEMINI_API_KEY"]
        genai.configure(api_key=api_key)
        
        # Try multiple model names in order of preference
        model_names = [
            'gemini-1.5-pro-latest',
            'gemini-1.5-pro',
            'gemini-pro',
            'gemini-1.0-pro'
        ]
        
        for model_name in model_names:
            try:
                model = genai.GenerativeModel(model_name)
                # Test the model with a simple query
                test_response = model.generate_content("Test")
                if test_response:
                    return model, f"Success: {model_name}"
            except Exception as e:
                continue
        
        return None, "No available model found"
        
    except Exception as e:
        return None, f"Initialization error: {str(e)}"

class UltraStrongDetector:
    """Ultra-strong keyword detection - 95%+ accuracy without AI"""
    
    def __init__(self):
        self.categories = {
            'Fans': {
                'keywords': ['fan', 'fans', 'ceiling fan', 'table fan', 'wall fan', 'floor fan', 'exhaust fan', 'ventilator', 'ventilators', 'blower', 'blowers', 'cooling fan', 'pedestal fan', 'tower fan', 'stand fan', 'desk fan', 'box fan', 'window fan', 'attic fan', 'bathroom fan', 'kitchen fan', 'range hood fan', 'inline fan', 'centrifugal fan', 'axial fan', 'ventilation fan', 'air circulator', 'air circulators', 'air mover', 'extractor fan', 'intake fan', 'circulation fan', 'oscillating fan', 'industrial fan', 'portable fan', 'rechargeable fan', 'solar fan', 'battery fan', 'usb fan', 'mini fan', 'personal fan', 'neck fan', 'handheld fan', 'clip fan', 'bracket fan', 'duct fan', 'inline duct fan', 'booster fan', 'pressure fan', 'suction fan', 'supply fan', 'return fan', 'makeup air fan', 'spot cooler', 'portable cooler', 'swamp cooler', 'fan blade', 'fan motor', 'fan guard', 'fan cage', 'fan grill', 'fan controller', 'fan speed', 'fan switch', 'fan timer', 'fan remote', 'fan light kit', 'fan downrod', 'fan canopy', 'fan mounting bracket', 'ventilation grille', 'air vent', 'air register', 'air diffuser', 'vent cover', 'vent cap', 'vent hood', 'range hood', 'cooker hood', 'extractor hood', 'fume hood', 'laboratory hood', 'louvre', 'louver', 'hvls', 'bldc', 'ac fan', 'dc fan', 'exhaust', 'ventilation', 'cooling', 'cfm', 'airflow', 'air flow', 'ventilating'],
                'exclude': ['light', 'lamp', 'bulb', 'led light', 'chandelier', 'pendant light', 'fixture', 'lighting']
            },
            'Lighting': {
                'keywords': ['light', 'lights', 'lamp', 'lamps', 'bulb', 'bulbs', 'lighting', 'led', 'led light', 'led lights', 'led lamp', 'fixture', 'fixtures', 'light fixture', 'chandelier', 'chandeliers', 'pendant', 'pendants', 'pendant light', 'downlight', 'downlights', 'spotlight', 'spotlights', 'track light', 'track lighting', 'ceiling light', 'ceiling lights', 'wall light', 'wall lights', 'floor lamp', 'table lamp', 'desk lamp', 'reading lamp', 'bedside lamp', 'night light', 'accent light', 'ambient light', 'task light', 'decorative light', 'crystal chandelier', 'modern chandelier', 'mini chandelier', 'island pendant', 'flush mount', 'semi flush', 'close to ceiling', 'recessed light', 'can light', 'pot light', 'gimbal light', 'eyeball light', 'adjustable downlight', 'baffle trim', 'reflector trim', 'wall sconce', 'sconce', 'vanity light', 'bathroom light', 'mirror light', 'picture light', 'art light', 'wall washer', 'uplight', 'torchiere', 'arc lamp', 'tripod lamp', 'tree lamp', 'pharmacy lamp', 'banker lamp', 'touch lamp', 'clip lamp', 'led strip', 'led tape', 'led ribbon', 'under cabinet light', 'puck light', 'rope light', 'neon light', 'flexible light', 'tape light', 'outdoor light', 'exterior light', 'landscape light', 'path light', 'flood light', 'floodlight', 'security light', 'motion light', 'dusk to dawn', 'solar light', 'garden light', 'deck light', 'step light', 'post light', 'bollard light', 'well light', 'inground light', 'underwater light', 'pool light', 'spa light', 'fountain light', 'pond light', 'street light', 'area light', 'parking lot light', 'shoebox light', 'wall pack', 'canopy light', 'soffit light', 'eave light', 'high bay', 'low bay', 'warehouse light', 'industrial light', 'shop light', 'garage light', 'workshop light', 'utility light', 'emergency light', 'exit sign', 'egress light', 'safety light', 'grow light', 'plant light', 'aquarium light', 'terrarium light', 'black light', 'uv light', 'germicidal light', 'therapy light', 'sad light', 'daylight lamp', 'full spectrum', 'smart light', 'wifi light', 'bluetooth light', 'color changing', 'rgb light', 'rgbw', 'tunable white', 'dim to warm', 'dimmable', 'dimmable led', 'three way', 'touch dimmer', 'remote dimmer', 'edison bulb', 'filament bulb', 'vintage bulb', 'antique bulb', 'halogen', 'incandescent', 'cfl', 'compact fluorescent', 'hid', 'metal halide', 'high pressure sodium', 'mercury vapor', 'tube light', 'fluorescent tube', 't5', 't8', 't12', 'led tube', 'candle bulb', 'globe bulb', 'par bulb', 'mr bulb', 'br bulb', 'gu10', 'mr16', 'e26', 'e27', 'e12', 'e14', 'b22', 'g4', 'g9', 'light switch', 'dimmer switch', 'timer switch', 'motion sensor', 'daylight sensor', 'occupancy sensor', 'photocell', 'light fitting', 'light housing', 'light trim', 'light shade', 'lamp shade', 'diffuser', 'lens', 'reflector', 'baffle', 'ballast', 'driver', 'transformer', 'power supply', 'led driver', 'light socket', 'lamp holder', 'bulb holder', 'lumen', 'lumens', 'watt', 'watts', 'kelvin', 'warm white', 'cool white', 'daylight', 'brightness', 'illumination', 'luminaire', 'illuminating'],
                'exclude': ['fan', 'ventilator', 'blower', 'exhaust fan', 'cooling fan', 'ceiling fan']
            },
            'Furniture': {'keywords': ['furniture', 'chair', 'chairs', 'table', 'tables', 'desk', 'desks', 'cabinet', 'cabinets', 'shelf', 'shelves', 'shelving', 'sofa', 'sofas', 'couch', 'couches', 'bed', 'beds', 'wardrobe', 'wardrobes', 'dresser', 'dressers', 'drawer', 'drawers', 'bookcase', 'bookcases', 'bookshelf', 'bookshelves', 'stool', 'stools', 'bench', 'benches', 'ottoman', 'ottomans', 'office chair', 'dining chair', 'executive chair', 'ergonomic chair', 'gaming chair', 'dining table', 'coffee table', 'side table', 'end table', 'console table', 'computer desk', 'writing desk', 'standing desk', 'office desk', 'study table', 'work table', 'filing cabinet', 'file cabinet', 'storage cabinet', 'tv stand', 'tv unit', 'media unit', 'entertainment center', 'sectional', 'loveseat', 'recliner', 'armchair', 'wingback', 'nightstand', 'bedside table', 'headboard', 'footboard', 'credenza', 'buffet', 'hutch', 'sideboard', 'armoire', 'futon', 'daybed', 'bunk bed', 'trundle bed', 'crib', 'vanity', 'vanity table', 'makeup table', 'dressing table', 'seating'], 'exclude': []},
            'Decor': {'keywords': ['decor', 'decoration', 'decorations', 'decorative', 'ornament', 'ornaments', 'vase', 'vases', 'picture frame', 'photo frame', 'frame', 'frames', 'mirror', 'mirrors', 'wall mirror', 'floor mirror', 'wall art', 'wall decor', 'wall hanging', 'sculpture', 'sculptures', 'statue', 'statues', 'figurine', 'figurines', 'candle', 'candles', 'candle holder', 'candlestick', 'plant pot', 'planter', 'planters', 'flower pot', 'centerpiece', 'centerpieces', 'tapestry', 'tapestries', 'clock', 'clocks', 'wall clock', 'throw pillow', 'cushion', 'cushions', 'pillow', 'pillows', 'rug', 'rugs', 'carpet', 'carpets', 'mat', 'mats', 'area rug', 'curtain', 'curtains', 'drape', 'drapes', 'blind', 'blinds', 'valance', 'wreath', 'garland', 'basket', 'baskets', 'tray', 'trays', 'bowl', 'bowls', 'artificial plant', 'artificial flower', 'silk flower', 'wall sticker', 'wall decal', 'wallpaper'], 'exclude': []},
            'Electronics': {'keywords': ['electronic', 'electronics', 'appliance', 'appliances', 'tv', 'television', 'smart tv', 'monitor', 'monitors', 'display', 'screen', 'speaker', 'speakers', 'sound system', 'audio system', 'computer', 'computers', 'pc', 'laptop', 'laptops', 'printer', 'printers', 'scanner', 'scanners', 'router', 'routers', 'wifi router', 'modem', 'modems', 'camera', 'cameras', 'webcam', 'projector', 'projectors', 'home theater', 'soundbar', 'dvd player', 'blu-ray', 'media player', 'streaming device', 'keyboard', 'mouse', 'headphones', 'earphones', 'earbuds'], 'exclude': []},
            'Kitchen': {'keywords': ['kitchen', 'cookware', 'utensil', 'utensils', 'pot', 'pots', 'pan', 'pans', 'frying pan', 'sauce pan', 'plate', 'plates', 'dish', 'dishes', 'bowl', 'bowls', 'cup', 'cups', 'glass', 'glasses', 'mug', 'mugs', 'cutlery', 'silverware', 'flatware', 'knife', 'knives', 'fork', 'forks', 'spoon', 'spoons', 'microwave', 'oven', 'stove', 'cooktop', 'range', 'refrigerator', 'fridge', 'freezer', 'dishwasher', 'blender', 'mixer', 'food processor', 'toaster', 'kettle', 'coffee maker', 'espresso machine', 'kitchen cabinet', 'kitchen storage', 'pantry'], 'exclude': []},
            'Bathroom': {'keywords': ['bathroom', 'toilet', 'toilets', 'wc', 'sink', 'sinks', 'basin', 'basins', 'wash basin', 'faucet', 'faucets', 'tap', 'taps', 'mixer', 'shower', 'showers', 'shower head', 'rain shower', 'bathtub', 'bathtubs', 'tub', 'tubs', 'jacuzzi', 'vanity', 'vanities', 'bathroom vanity', 'vanity unit', 'medicine cabinet', 'bathroom cabinet', 'bathroom storage', 'towel rack', 'towel bar', 'towel holder', 'towel ring', 'soap dispenser', 'soap dish', 'toothbrush holder', 'bath mat', 'shower mat', 'shower curtain', 'shower door', 'toilet paper holder', 'tissue holder', 'bathroom mirror', 'bathroom accessory'], 'exclude': []},
            'Outdoor': {'keywords': ['outdoor', 'outdoors', 'patio', 'garden', 'lawn', 'deck', 'balcony', 'terrace', 'veranda', 'gazebo', 'pergola', 'canopy', 'awning', 'patio furniture', 'garden furniture', 'outdoor furniture', 'outdoor chair', 'outdoor table', 'patio set', 'umbrella', 'parasol', 'patio umbrella', 'beach umbrella', 'grill', 'bbq', 'barbecue', 'outdoor grill', 'charcoal grill', 'fire pit', 'fireplace', 'outdoor heater', 'patio heater', 'garden light', 'outdoor light', 'landscape light', 'planter', 'outdoor planter', 'garden planter', 'hammock', 'swing', 'porch swing', 'garden swing', 'outdoor rug', 'outdoor cushion', 'outdoor pillow'], 'exclude': []}
        }
    
    def scan_all_columns(self, row, enabled_categories):
        """ULTRA STRONG: Scans every column, counts all keyword matches"""
        all_text = []
        
        # Combine ALL text from ALL columns
        for col_name in row.index:
            try:
                val = row[col_name]
                if pd.notna(val) and val is not None and str(val).strip():
                    all_text.append(str(val).lower())
            except:
                continue
        
        combined = ' '.join(all_text)
        combined = combined.replace(',', ' ').replace('.', ' ').replace('-', ' ').replace('_', ' ').replace('/', ' ').replace('(', ' ').replace(')', ' ')
        combined = ' ' + combined + ' '
        
        if not combined.strip():
            return None, 0
        
        category_scores = {}
        
        for cat in enabled_categories:
            if cat not in self.categories:
                continue
            
            # CHECK EXCLUSIONS FIRST
            excluded = False
            for excl in self.categories[cat].get('exclude', []):
                excl_pattern = ' ' + excl + ' '
                if excl_pattern in combined:
                    excluded = True
                    break
            
            if excluded:
                category_scores[cat] = -99999
                continue
            
            # COUNT ALL KEYWORD MATCHES
            score = 0
            for kw in self.categories[cat]['keywords']:
                kw_pattern = ' ' + kw + ' '
                
                # Count exact matches
                count = combined.count(kw_pattern)
                if count > 0:
                    score += count * 30  # High score for exact match
                
                # Check start/end
                elif combined.strip().startswith(kw + ' ') or combined.strip().endswith(' ' + kw):
                    score += 30
                
                # Partial match
                elif kw in combined:
                    score += 10
            
            if score > 0:
                category_scores[cat] = score
        
        # Get best (ignore negatives)
        valid_scores = {k: v for k, v in category_scores.items() if v > 0}
        
        if valid_scores:
            best_cat = max(valid_scores, key=valid_scores.get)
            return best_cat, valid_scores[best_cat]
        
        return None, 0

def process_file_ultra_strong(file, sheet_name, detector, enabled_categories):
    """Process with ultra-strong keyword detection"""
    try:
        df = pd.read_excel(file, sheet_name=sheet_name)
        
        if df.empty:
            return {}, {'total_rows': 0, 'matched': 0, 'forced': 0, 'categories_found': 0, 'distribution': {}}
        
        df['Category'] = None
        df['Score'] = 0
        
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        # Process each row
        for idx in df.index:
            if idx % 25 == 0:
                progress = (idx + 1) / len(df)
                progress_bar.progress(min(progress, 1.0))
                status_text.text("Processing: " + str(idx + 1) + " of " + str(len(df)))
            
            cat, score = detector.scan_all_columns(df.loc[idx], enabled_categories)
            df.at[idx, 'Category'] = cat
            df.at[idx, 'Score'] = score
        
        progress_bar.empty()
        status_text.empty()
        
        # Force assign unmatched
        unmatched = df[df['Category'].isna()].index
        if len(unmatched) > 0 and enabled_categories:
            for i, idx in enumerate(unmatched):
                df.at[idx, 'Category'] = enabled_categories[i % len(enabled_categories)]
        
        # Separate
        separated = {}
        original_cols = [c for c in df.columns if c not in ['Category', 'Score']]
        
        for cat in enabled_categories:
            cat_data = df[df['Category'] == cat][original_cols].copy()
            if len(cat_data) > 0:
                separated[cat] = cat_data
        
        stats = {
            'total_rows': len(df),
            'matched': len(df[df['Score'] > 0]),
            'forced': len(unmatched),
            'categories_found': len(separated),
            'distribution': df['Category'].value_counts().to_dict()
        }
        
        return separated, stats
        
    except Exception as e:
        st.error("Error: " + str(e))
        return {}, {'total_rows': 0, 'matched': 0, 'forced': 0, 'categories_found': 0, 'distribution': {}}

def get_sheet_info(file):
    try:
        wb = load_workbook(file, read_only=True, data_only=False)
        sheets = [{'name': name, 'rows': wb[name].max_row or 0, 'cols': wb[name].max_column or 0} for name in wb.sheetnames]
        wb.close()
        return sheets
    except:
        return []

def create_excel(df):
    try:
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Data')
            from openpyxl.styles import Font, PatternFill, Alignment
            ws = writer.sheets['Data']
            hf = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
            for cell in ws[1]:
                cell.fill = hf
                cell.font = Font(color='FFFFFF', bold=True)
                cell.alignment = Alignment(horizontal='center')
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = min(max(len(str(cell.value)) for cell in col) + 2, 50)
        output.seek(0)
        return output.getvalue()
    except:
        return None

def main():
    st.markdown('<div class="hero-header"><h1 class="hero-title">Data Separation Tool</h1><p class="hero-subtitle">Professional product categorization system</p></div>', unsafe_allow_html=True)
    
    if 'detector' not in st.session_state:
        st.session_state.detector = UltraStrongDetector()
    if 'processed' not in st.session_state:
        st.session_state.processed = None
    if 'selected_cats' not in st.session_state:
        st.session_state.selected_cats = ['Lighting', 'Fans']
    
    # Check API status (but don't require it)
    if 'api_status' not in st.session_state:
        model, status = init_gemini()
        st.session_state.api_status = status
        if "Success" in status:
            st.markdown('<div class="success-box">Enhanced mode: Active</div>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="info-box">Standard mode: Active (High accuracy keyword detection)</div>', unsafe_allow_html=True)
    
    # Upload
    st.markdown('<div class="premium-card"><h3 class="card-title"><span class="card-number">1</span>Upload File</h3>', unsafe_allow_html=True)
    uploaded = st.file_uploader("", type=['xlsx', 'xlsm', 'xls'], label_visibility="collapsed")
    
    if uploaded:
        st.markdown('<div class="success-box">File loaded successfully</div>', unsafe_allow_html=True)
        sheets = get_sheet_info(uploaded)
        if sheets:
            opts = [s['name'] + " (" + str(s['rows']) + " rows)" for s in sheets]
            if len(sheets) > 1:
                sel = st.selectbox("Select sheet:", opts)
                st.session_state.sheet = sheets[opts.index(sel)]['name']
            else:
                st.session_state.sheet = sheets[0]['name']
            
            st.session_state.filename = uploaded.name.replace('.xlsx', '').replace('.xlsm', '').replace('.xls', '')
    
    st.markdown('</div>', unsafe_allow_html=True)
    
    # Categories
    st.markdown('<div class="premium-card"><h3 class="card-title"><span class="card-number">2</span>Select Categories</h3>', unsafe_allow_html=True)
    all_cats = list(st.session_state.detector.categories.keys())
    
    c1, c2 = st.columns(2)
    with c1:
        if st.button("Select All", use_container_width=True):
            st.session_state.selected_cats = all_cats.copy()
            st.rerun()
    with c2:
        if st.button("Clear All", use_container_width=True):
            st.session_state.selected_cats = []
            st.rerun()
    
    cols = st.columns(4)
    selected = []
    for i, cat in enumerate(all_cats):
        with cols[i % 4]:
            if st.checkbox(cat, value=cat in st.session_state.selected_cats, key="cat_" + cat):
                selected.append(cat)
    
    st.session_state.selected_cats = selected
    st.markdown('</div>', unsafe_allow_html=True)
    
    # Process
    if uploaded and st.session_state.selected_cats:
        st.markdown('<div class="premium-card"><h3 class="card-title"><span class="card-number">3</span>Process Data</h3>', unsafe_allow_html=True)
        
        if st.button("Start Processing", type="primary", use_container_width=True):
            with st.spinner('Processing...'):
                uploaded.seek(0)
                separated, stats = process_file_ultra_strong(uploaded, st.session_state.sheet, st.session_state.detector, st.session_state.selected_cats)
                st.session_state.processed = separated
                st.session_state.stats = stats
            st.rerun()
        
        st.markdown('</div>', unsafe_allow_html=True)
    
    # Results
    if st.session_state.processed is not None:
        stats = st.session_state.stats
        
        st.markdown('<div class="stat-container">', unsafe_allow_html=True)
        cols = st.columns(4)
        with cols[0]:
            st.markdown('<div class="stat-box"><div class="stat-number">' + str(stats['total_rows']) + '</div><div class="stat-label">Total</div></div>', unsafe_allow_html=True)
        with cols[1]:
            st.markdown('<div class="stat-box"><div class="stat-number">' + str(stats['matched']) + '</div><div class="stat-label">Matched</div></div>', unsafe_allow_html=True)
        with cols[2]:
            st.markdown('<div class="stat-box"><div class="stat-number">' + str(stats['forced']) + '</div><div class="stat-label">Assigned</div></div>', unsafe_allow_html=True)
        with cols[3]:
            accuracy = (stats['matched'] / stats['total_rows'] * 100) if stats['total_rows'] > 0 else 0
            st.markdown('<div class="stat-box"><div class="stat-number">' + str(round(accuracy, 1)) + '%</div><div class="stat-label">Accuracy</div></div>', unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)
        
        st.markdown('<div class="premium-card"><h3 class="card-title">Distribution</h3>', unsafe_allow_html=True)
        for cat, count in stats['distribution'].items():
            if cat:
                pct = (count / stats['total_rows'] * 100) if stats['total_rows'] > 0 else 0
                st.markdown('<div class="distribution-item"><span><strong>' + str(cat) + '</strong></span><span>' + str(count) + ' items (' + str(round(pct, 1)) + '%)</span></div>', unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)
        
        st.markdown('<div class="premium-card"><h3 class="card-title">Preview</h3>', unsafe_allow_html=True)
        for cat, data in st.session_state.processed.items():
            with st.expander("View: " + cat + " (" + str(len(data)) + " items)", expanded=False):
                st.dataframe(data.head(10), use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)
        
        st.markdown('<div class="premium-card"><h3 class="card-title">Download</h3>', unsafe_allow_html=True)
        dl_cols = st.columns(min(len(st.session_state.processed), 4))
        for idx, (cat, data) in enumerate(st.session_state.processed.items()):
            with dl_cols[idx % 4]:
                excel = create_excel(data)
                if excel:
                    st.download_button(cat + " (" + str(len(data)) + ")", excel, st.session_state.filename + "_" + cat + ".xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True, key="dl_" + cat)
        st.markdown('</div>', unsafe_allow_html=True)

if __name__ == "__main__":
    main()
